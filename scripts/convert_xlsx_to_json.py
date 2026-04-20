#!/usr/bin/env python3
"""
조특법 체크리스트 엑셀 → JSON 변환기

입력 (data/*.xlsx):
  - 조특법.xlsx                                  (시트: 혜택내용)
  - 조특법_127조_중복지원배제_조견표.xlsx        (시트: 프로그램용_통합매트릭스, 조문_구조요약)
  - 조특법_128조_추계과세감면배제_조견표.xlsx
  - 조특법_132조_최저한세_조견표.xlsx
  - 농어촌특별세_조특법감면_과세비과세_조견표.xlsx (시트: 프로그램용_통합DB)

출력 (data/*.json):
  - 조특법_공제목록.json        : 메인 공제 카드 리스트
  - 조특법_조문상세.json         : 조문별 농특세·128·132 통합
  - 조특법_중복배제매트릭스.json : 127조 A↔B 페어
  - 조특법_메타.json             : groupMap, 생성일자 등
"""
from __future__ import annotations
import json
import re
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

XLSX_MAIN       = DATA / "조특법.xlsx"
XLSX_127        = DATA / "조특법_127조_중복지원배제_조견표.xlsx"
XLSX_128        = DATA / "조특법_128조_추계과세감면배제_조견표.xlsx"
XLSX_132        = DATA / "조특법_132조_최저한세_조견표.xlsx"
XLSX_NONGTEUK   = DATA / "농어촌특별세_조특법감면_과세비과세_조견표.xlsx"

# ── 유틸 ────────────────────────────────────────────────────────────────────

CIRCLED = "①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳"

def circled_to_num(s: str) -> str:
    """①②③ → 1, 2, 3"""
    for i, ch in enumerate(CIRCLED, 1):
        s = s.replace(ch, f"({i})")
    return s

def normalize_article_key(s) -> str:
    """조문 표기를 캐노니컬 키로 변환.

    예시:
      '조특법10조'          → '10'
      '조특법 30조'         → '30'
      '조특법12조의2'       → '12-2'
      '조특법104조의24 1항' → '104-24.1'
      '제6조'               → '6'
      '제12조의2'           → '12-2'
      '제8조의3'            → '8-3'
      '§8의3③'             → '8-3.3'
      '§24'                 → '24'
      '§19①'               → '19.1'
    """
    if s is None: return ""
    t = str(s).strip()
    if not t: return ""

    # 원문자(circled) → (N)
    t = circled_to_num(t)

    # 접두어 제거
    t = t.replace("조특법", "").replace("제", "").replace("§", "").strip()

    # "104조의24 1항" 또는 "104조의24 제1항" → "104-24.1"
    # 먼저 항(hang) 분리
    hang = ""
    m = re.search(r"(?:\s+|제)(\d+)\s*항", t)
    if m:
        hang = "." + m.group(1)
        t = t[:m.start()] + t[m.end():]
    # "(N)" 형태(원래 원문자) 분리
    m = re.search(r"\((\d+)\)", t)
    if m:
        hang = "." + m.group(1) if not hang else hang
        t = t[:m.start()] + t[m.end():]

    # "1~2항" 같은 레인지는 첫 숫자만 사용
    m = re.match(r".*?(\d+)\s*[~\-]\s*\d+\s*항", t)
    if m and not hang:
        hang = "." + m.group(1)
        t = re.sub(r"\s*\d+\s*[~\-]\s*\d+\s*항.*$", "", t)

    # "조" 제거, "의" → "-"
    t = t.replace("조", "").replace("의", "-").strip()
    t = re.sub(r"\s+", "", t)

    return t + hang


def split_lines(s) -> list[str]:
    """개행·파이프·쉼표 분리해 빈 줄 제거."""
    if s is None: return []
    out = []
    for chunk in re.split(r"[\n|]", str(s)):
        c = chunk.strip().strip(",")
        if c: out.append(c)
    return out


def split_csv(s) -> list[str]:
    if s is None: return []
    return [x.strip() for x in re.split(r"[,\n]", str(s)) if x.strip()]


def cell(ws, r, c):
    v = ws.cell(r, c).value
    if v is None: return ""
    return str(v).strip()


def load_wb(path: Path):
    if not path.exists():
        print(f"[경고] 파일 없음: {path.name}", file=sys.stderr)
        return None
    return openpyxl.load_workbook(path, data_only=True)


# ── 1. 혜택내용 (메인 공제 목록) ──────────────────────────────────────────────

def build_deductions() -> list[dict]:
    wb = load_wb(XLSX_MAIN)
    if wb is None: return []
    ws = wb["혜택내용"]
    out = []
    seen = {}
    for r in range(2, ws.max_row + 1):
        id_ = cell(ws, r, 1)
        article = cell(ws, r, 2)
        title = cell(ws, r, 3)
        if not id_ and not title: continue
        # 같은 id 중복 시: 비어있지 않은 값이 있는 행으로 병합
        entry = {
            "id": id_,
            "article": article,
            "article_key": normalize_article_key(article),
            "title": title,
            "tags": split_csv(cell(ws, r, 4)),
            "reqs": split_lines(cell(ws, r, 5)),
            "management": cell(ws, r, 6),
            "exclusion_codes": [c for c in split_csv(cell(ws, r, 7)) if c and c != "NONE"],
            "agri_tax":    cell(ws, r, 8),
            "min_tax":     cell(ws, r, 9),
            "deemed_depr": cell(ws, r, 10),
            "carry_over":  cell(ws, r, 11),
        }
        if id_ in seen:
            # 병합: 기존 빈 필드를 새 값으로 채움
            prev = seen[id_]
            for k, v in entry.items():
                if not prev.get(k) and v:
                    prev[k] = v
        else:
            seen[id_] = entry
            out.append(entry)
    return out


# ── 2. 농특세 통합DB + 128·132 조문별 상세 ───────────────────────────────────

def build_article_detail() -> dict:
    """article_key → {title, agri_tax, estimation_128, min_tax_132}"""
    detail: dict = {}

    # 2-1. 농특세 통합DB (주 레코드)
    wb = load_wb(XLSX_NONGTEUK)
    if wb is not None and "프로그램용_통합DB" in wb.sheetnames:
        ws = wb["프로그램용_통합DB"]
        # 헤더는 row 3: 조문 | 감면·공제 명칭 | 농특세 과세여부 | 근거 | 세율 | 128조 | 132조
        for r in range(4, ws.max_row + 1):
            art_raw = cell(ws, r, 1)
            if not art_raw: continue
            key = normalize_article_key(art_raw)
            if not key: continue
            # 첫 등록만 채우고, 나머지는 병합 노트로만
            rec = detail.setdefault(key, {
                "article_display": art_raw,
                "title": "",
                "agri_tax": {},
                "estimation_128": {},
                "min_tax_132": {},
            })
            if not rec["title"]:
                rec["title"] = cell(ws, r, 2)
            rec["agri_tax"] = {
                "status": cell(ws, r, 3),   # 과세 / 비과세
                "basis":  cell(ws, r, 4),
                "rate":   cell(ws, r, 5),
            }
            rec["estimation_128"]["flag"] = cell(ws, r, 6)  # O / -
            rec["min_tax_132"]["flag"]    = cell(ws, r, 7)  # O / - / O(100%연도제외)

    # 2-2. 128조 배제항목 — 어떤 항에서 배제되는지 수집
    wb = load_wb(XLSX_128)
    if wb is not None:
        for sheet_name, para_label in [
            ("제1항_감면배제항목",  "제1항"),
            ("제2_4항_감면배제항목", "제2~4항"),
        ]:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            for r in range(4, ws.max_row + 1):
                art_raw = cell(ws, r, 2)
                name    = cell(ws, r, 3)
                note    = cell(ws, r, 4)
                if not art_raw: continue
                key = normalize_article_key(art_raw)
                if not key: continue
                rec = detail.setdefault(key, {
                    "article_display": art_raw,
                    "title": name,
                    "agri_tax": {}, "estimation_128": {}, "min_tax_132": {},
                })
                rec["estimation_128"].setdefault("paragraphs", []).append({
                    "para": para_label,
                    "name": name,
                    "note": note,
                })

    # 2-3. 132조 최저한세 대상 — 법인세·소득세 구분
    wb = load_wb(XLSX_132)
    if wb is not None:
        for sheet_name, tax_kind in [
            ("법인세_최저한세대상", "법인세"),
            ("소득세_최저한세대상", "소득세"),
            ("최저한세_제외조문",   "제외"),
        ]:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            # 헤더 row 3, 조문 열은 col 2
            art_col = 2
            name_col = 3
            note_col = 4
            for r in range(4, ws.max_row + 1):
                art_raw = cell(ws, r, art_col)
                if not art_raw: continue
                key = normalize_article_key(art_raw)
                if not key: continue
                rec = detail.setdefault(key, {
                    "article_display": art_raw,
                    "title": cell(ws, r, name_col),
                    "agri_tax": {}, "estimation_128": {}, "min_tax_132": {},
                })
                rec["min_tax_132"].setdefault("applies", []).append({
                    "kind": tax_kind,     # 법인세 / 소득세 / 제외
                    "name": cell(ws, r, name_col),
                    "note": cell(ws, r, note_col),
                })

    return detail


# ── 3. 127조 중복배제 매트릭스 ───────────────────────────────────────────────

def build_127_matrix() -> list[dict]:
    wb = load_wb(XLSX_127)
    if wb is None or "프로그램용_통합매트릭스" not in wb.sheetnames:
        return []
    ws = wb["프로그램용_통합매트릭스"]
    out = []
    for r in range(5, ws.max_row + 1):
        a = cell(ws, r, 1)
        if not a: continue
        out.append({
            "a_article":  a,
            "a_key":      normalize_article_key(a),
            "a_name":     cell(ws, r, 2),
            "b_article":  cell(ws, r, 3),
            "b_key":      normalize_article_key(cell(ws, r, 3)),
            "b_name":     cell(ws, r, 4),
            "relation":   cell(ws, r, 5),   # 예: '택1 (동일 투자자산)'
            "basis":      cell(ws, r, 6),   # 예: '§127②'
        })
    return out


# ── 4. 127조 조문 구조요약 (모달에서 보여줄 항별 설명) ──────────────────────

def build_127_structure() -> list[dict]:
    wb = load_wb(XLSX_127)
    if wb is None or "조문_구조요약" not in wb.sheetnames:
        return []
    ws = wb["조문_구조요약"]
    out = []
    for r in range(5, ws.max_row + 1):
        para = cell(ws, r, 1)
        if not para: continue
        out.append({
            "para":    para,              # 제1항 / 제2항 ...
            "type":    cell(ws, r, 2),
            "title":   cell(ws, r, 3),
            "scope":   cell(ws, r, 4),
            "content": cell(ws, r, 5),
        })
    return out


# ── 5. 메인 ──────────────────────────────────────────────────────────────────

def main():
    deductions = build_deductions()
    detail     = build_article_detail()
    matrix     = build_127_matrix()
    structure  = build_127_structure()

    # 공제목록 → 매칭되는 127조 페어 id 리스트 주입 (사전계산)
    key_to_pairs: dict = {}
    for i, p in enumerate(matrix):
        key_to_pairs.setdefault(p["a_key"], []).append(i)
        key_to_pairs.setdefault(p["b_key"], []).append(i)
    for d in deductions:
        d["conflict_pairs"] = key_to_pairs.get(d["article_key"], [])

    # 메타
    kst = timezone(timedelta(hours=9))
    meta = {
        "generated_at": datetime.now(kst).isoformat(timespec="seconds"),
        "group_map": {
            "G2": "제127조 제2항 (세액공제 간 동일 투자자산·과세연도 택1)",
            "G3": "제127조 제3항 (세액감면 vs 세액공제 중복배제)",
            "G4": "제127조 제4항 (감면·공제 매트릭스)",
            "G5": "제127조 제5항 (세액감면 간 중복 택1)",
        },
        "counts": {
            "deductions": len(deductions),
            "article_detail": len(detail),
            "127_matrix_pairs": len(matrix),
            "127_structure_paragraphs": len(structure),
        },
    }

    # 출력
    outputs = {
        "조특법_공제목록.json":        deductions,
        "조특법_조문상세.json":         detail,
        "조특법_중복배제매트릭스.json": {"structure": structure, "pairs": matrix},
        "조특법_메타.json":             meta,
    }
    for name, data in outputs.items():
        path = DATA / name
        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        size = path.stat().st_size
        print(f"[OK] {name}  ({size:,} bytes)")

    print()
    print(f"  공제: {meta['counts']['deductions']}개")
    print(f"  조문 상세: {meta['counts']['article_detail']}개")
    print(f"  중복배제 페어: {meta['counts']['127_matrix_pairs']}개")
    print(f"  127조 항 구조: {meta['counts']['127_structure_paragraphs']}개")


if __name__ == "__main__":
    main()
